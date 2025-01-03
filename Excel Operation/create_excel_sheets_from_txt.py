import os

from openpyxl import Workbook

# Define the working file path
txt_file_path = r"D:\YK\Downloads\sheets.txt"
current_directory = os.path.dirname(os.path.abspath(txt_file_path))
excel_file_name = 'output.xlsx'
excel_file_path = os.path.join(current_directory, excel_file_name)

# Read sheet names from the .txt file
with open(txt_file_path, 'r') as file:
    sheet_names = [line.strip() for line in file.readlines()]

# Create a new workbook and the Summary sheet
wb = Workbook()
summary_sheet = wb.active
summary_sheet.title = "Summary"

# Add the specified sheets to the workbook
for sheet_name in sheet_names:
    wb.create_sheet(title=sheet_name)

# Populate the Summary sheet with sheet names and links
summary_sheet.cell(row=1, column=1, value="Sheet Name")
summary_sheet.cell(row=1, column=2, value="Link")

for idx, sheet_name in enumerate(sheet_names, start=2):
    summary_sheet.cell(row=idx, column=1, value=sheet_name)
    summary_sheet.cell(row=idx, column=2, value=f"=HYPERLINK(\"#{sheet_name}!A1\", \"Go to {sheet_name}\")")

# Save the workbook in the same directory as the .txt file
wb.save(excel_file_path)

print(f"Excel file created successfully at {excel_file_path}!")
